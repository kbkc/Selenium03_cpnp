from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from dicttoxml import dicttoxml
from xmltodict import parse, unparse
#from main import ini_data



fn_session_data = 'session_data.xml'


kw_list = ('url', 'click', 'input','execute_script','get_text')

chromedriver_name = './chromedriver'                   # in linux run app
if os.name == 'nt': chromedriver_name = 'chromedriver' # in windows run app

def write_session_data(dict):
    out = dicttoxml(dict, custom_root='session_data', attr_type=False)
    with open(fn_session_data, 'wb') as f:
        f.write(out)
        f.close()


def read_session_data():
    with open(fn_session_data, 'rb') as f:
        xmld = f.read()
        d = parse(xmld)
        f.close()
    return {'id': d['session_data']['id'], 'url':  d['session_data']['url']}
    

def run_rule(driver,rules_list):
    lr = parce_rules_list(rules_list)

    for subls in lr:
        run_subrule(driver, subls)
        time.sleep(3)

def run_rule_dbl(driver,rules_list,ini_data):
    lr = parce_rules_list(rules_list)

    for subls in lr:
        print(subls)
        run_subrule(driver, subls)
        #time.sleep(3)

def parce_rules_list(rl):
    print('parce rule: ', rl[0])
    lrules = []
    lsubrules = []
    f = True
    for i in range (1, len(rl)):   
        if  rl[i] in kw_list:
            if f is True:
                lsubrules.append(rl[i])
                f = False
            else:
                lrules.append(lsubrules)
                lsubrules = []
                lsubrules.append(rl[i])
                f = True
        else:
            lsubrules.append(rl[i])
    lrules.append(lsubrules)
    return lrules

# -------------------------------------------------------
#                 check_exists_by_xpath
#--------------------------------------------------------
def check_exists_by_xpath(webdriver,xpath):
    try:
        webdriver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True
# -------------------------------------------------------
#                 run_subrule
#--------------------------------------------------------
def run_subrule(driver, subls):

    getting_text = ''

    print('---------------------------------------------')
    print(f'rule: {subls}')

    if subls[0] != kw_list[0]:
        driver.implicitly_wait(5)
        if check_exists_by_xpath(driver,subls[1]) == False :
            print('   element not exist')
            return
    # url
    if subls[0] == kw_list[0]:

        elem = driver.get(subls[1])
    # click
    elif subls[0] == kw_list[1]:
        try:
            #el = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, subls[1])))
            elem = driver.find_element_by_xpath(subls[1])
            elem.click()
        except NoSuchElementException:
            print("Not found elements: ", subls[1]) 
        finally:
            pass

    # input 
    elif subls[0] == kw_list[2]:
        try:
            elem = driver.find_element_by_xpath(subls[1])
            elem.clear()
            elem.send_keys(subls[2].strip())
        except NoSuchElementException:
            print("Not found elements: ", subls[1], ', ',subls[2])
        finally:
            pass

    # execute_script
    elif subls[0] == kw_list[3]:
        try:
            elem = driver.find_element_by_xpath(subls[1])
            driver.execute_script("arguments[0].click();", elem)
        except NoSuchElementException:
            print("Not found elements: ", subls[1])  
        finally:
            pass   
            # execute_script
    elif subls[0] == kw_list[4]:
        try:
            elem = driver.find_element_by_xpath(subls[1])
            getting_text = elem.text
        except NoSuchElementException:
            print("Not found elements: ", subls[1])  
        finally:
            pass    
    return getting_text



def read_settings(fname, sheet_name):
    wb = openpyxl.load_workbook(fname)
    sh = wb.get_sheet_by_name(sheet_name)
    lst = []
    lst2=[]
    for row in sh.iter_rows():
        lst2=[]
        for cell in row:
            # 2020-08-04 for correct reading tables
            #if cell.value != None : lst2.append(str(cell.value))
            
            lst2.append(str(cell.value))
        lst.append(lst2)       
    wb.close()
    return(lst)


def read_product_table(fname, sheet_name,table_name):
    wb = openpyxl.load_workbook(fname)
    sh = wb.get_sheet_by_name(sheet_name)
    table = sh.tables[table_name]
    lst = []
    for x in sh[table.ref]:
        lst2=[]
        for y in x:
            #print(y.value)
            lst2.append(y.value)
        lst.append(lst2)
        
    '''
    lst = []
    lst2=[]
    for row in sh.iter_rows():
        lst2=[]
        for cell in row:
            # 2020-08-04 for correct reading tables
            #if cell.value != None : lst2.append(str(cell.value))
            
            lst2.append(str(cell.value))
        lst.append(lst2)   
    '''    
    wb.close()
    return(lst)



def open_url(surl):
    """opens browser with url 
    
    Arguments:
        surl {[type]} -- url-address
    """
    if not surl :
        print('url is needed')
        return 'error'
    if not surl.startswith('http')  :
        surl = 'http://' + surl
        print('no http')
    print('url = ', surl)
    driver = webdriver.Chrome(chromedriver_name)
    driver.get(surl)
    time.sleep(2)
    return driver




def close_driver(drv):
    drv.close()
    print('...closed')

# ############################################### #
# data for connect to session from other scripts  #
def show_data_connect_info(drv):
    url = drv.command_executor._url
    session_id = drv.session_id
    dict = {'id': session_id, 'url': url}
    write_session_data(dict)
    print(f'\n\t\t===\nurl = \'{url.strip()}\'\nsession_id = \'{session_id.strip()}\' \n \t\t===\n')
# ################################################ #







